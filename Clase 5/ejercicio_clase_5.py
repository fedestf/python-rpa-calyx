# Generar a partir del archivo "vgsales.csv" dos dataframes,
# uno deberá contener a los juegos cuyo género sea de Acción
# y el otro a los del género Sandbox. En ambos deben quedar
# solamente las columnas: title, genre, publisher, console.
import pandas as pd
from logger import logger_debug
from openpyxl import load_workbook

df = pd.read_csv(
    r'Clase 5\vgsales.csv')


df_accion = df[df["genre"] == "Action"][[
    "title", "genre", "publisher", "console"]]

df_sandbox = df[df["genre"] == "Sandbox"][[
    "title", "genre", "publisher", "console"]]


# Luego se debe obtener un tercer dataframe, a partir de los
# dos anteriores, que contenga aquellos juegos que esten presentes
# en PC (o alguna PS si no hay ocurrencias). Dicho dataframe
# debe ser guardado en formato excel.

df_accion_filtrado = df_accion.query('console=="PC" or console=="PS"')[
    ["title", "console", "genre", "publisher"]]
df_sandbox_filtrado = df_sandbox.query('console=="PC" or console=="PS"')[
    ["title", "console", "genre", "publisher"]]

df_concatenado = pd.concat([df_accion_filtrado, df_sandbox_filtrado])

df_concatenado.to_excel(
    r'Clase 5\datos_salida.xlsx', index=False)


# Una vez generado el excel, se deberá leer el mismo y loguear
# la información disponible en cada fila.

wb = load_workbook(
    r'Clase 5\datos_salida.xlsx')

sheet = wb['Sheet1']

for fila in sheet.iter_rows(min_col=sheet.min_column, max_col=sheet.max_column):

    logger_debug.debug("{}".format([celda.value for celda in fila]))
